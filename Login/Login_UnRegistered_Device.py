afsagsdsd
delattrvsd

dsvdsfdsfsd
git 
dsfdsfds
asfknsafmsa
sdfsdgsd
asfjasfnda


# Import necessary modules
from tkinter import simpledialog
import tkinter as tk
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.webdriver import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from tkinter import Tk, messagebox
from tkinter import Tk, filedialog
from openpyxl.styles import Font
import openpyxl
import time
import logging
import os
import getpass




# Logging configuration
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# Define constants
DEVICE_NAME = "D6HQXGN7BATODIPF"
APP_PACKAGE = "com.pnb.android.staging"
APP_ACTIVITY = "com.pnb.android.ui.splash.SplashActivity"



# Private Call-out
def show_message_box(message):
    root = Tk()
    root.withdraw()  
    messagebox.showinfo("Hold On!", message)
    root.destroy()  

show_message_box("Select Excel Database first!")


def get_excel_file_path():  
    root = Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    return file_path

EXCEL_FILE_PATH = get_excel_file_path()





# Define XPaths
XPATH_LOGIN_BUTTON = '//android.widget.Button[@resource-id="com.pnb.android.staging:id/btnLogin"]'
XPATH_USERNAME_FIELD = '//android.widget.EditText[@resource-id="com.pnb.android.staging:id/txtUsername"]'
XPATH_PASSWORD_FIELD = '//android.widget.EditText[@resource-id="com.pnb.android.staging:id/txtPassword"]'
XPATH_UNMASK_PASSWORD_BUTTON = '//android.widget.ImageButton[@content-desc="Show password"]'
XPATH_OAC_FIELD = '//android.widget.EditText[@resource-id="com.pnb.android.staging:id/txtOAC"]'
XPATH_CONFIRM_BUTTON = '//android.widget.Button[@resource-id="com.pnb.android.staging:id/btnConfirm"]'
XPATH_OTP_FIELD = '//android.widget.EditText[@resource-id="com.pnb.android.staging:id/txtOTP"]'







# Define desired capabilities
capabilities = {
    "platformName": "Android",
    "automationName": "UiAutomator2",
    "deviceName": DEVICE_NAME,
    "appPackage": APP_PACKAGE,
    "appActivity": APP_ACTIVITY,
    "noReset": False,
    "dontStopAppOnReset": True,
    "newCommandTimeout": 300,
}

appium_server_url = "http://127.0.0.1:4723"



# Global Call-out
def validate_elements_from_excel(driver, file_path, sheet_name, start_row=1, max_row=None):
    """Validate UI elements based on text and XPath from Excel."""
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Add header for result if not exists
        if sheet.cell(row=1, column=3).value != "Validation Result":
            sheet.cell(row=1, column=3).value = "Validation Result"

        # Define font styles
        red_font = Font(color="FF0000")  # Red
        green_font = Font(color="00B050")  # Optional: Green for matched text

        # Determine the maximum row dynamically
        last_row = max_row if max_row else sheet.max_row

        for row in sheet.iter_rows(min_row=start_row, max_row=last_row):  # Dynamic row range
            expected_text = row[0].value
            xpath = row[1].value

            if not xpath:
                logging.warning(f"Empty XPath in row {row[0].row}, skipping.")
                continue

            try:
                element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((AppiumBy.XPATH, xpath))
                )
                actual_text = element.text.strip()

                if actual_text == expected_text:
                    result = "✅ Match"
                    row[2].value = result
                    row[2].font = green_font
                    logging.info(f"[Row {row[0].row}] [Text Validation Pass]: '{actual_text}'")
                else:
                    result = f"❌ Text Mismatch (Expected: '{expected_text}', Found: '{actual_text}')"
                    row[2].value = result
                    row[2].font = red_font
                    logging.warning(f"[Row {row[0].row}] Text mismatch")

            except Exception as e:
                result = f"⛔ Element Not Found"
                row[2].value = result
                row[2].font = red_font
                logging.error(f"[Row {row[0].row}] Element not found using XPath: {xpath}")

        workbook.save(file_path)
        logging.info("Validation results saved to Excel.")

    except Exception as e:
        logging.error(f"Failed to validate elements from Excel: {str(e)}")


def scroll_down_and_find_element(by, value, max_attempts=5):
    attempts = 0
    while attempts < max_attempts:
        try:
            # Attempt to find the element directly without scrolling
            element = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((by, value))  # Wait for visibility, not just presence
            )
            print(f"Found element: Continue Execution")
            return element
        except Exception as e:
            print(f"Element not found in the visible area. Attempting to scroll down...")

            # Use UiScrollable to scroll down until the element is found
            try:
                # UiScrollable scroll command (scrollForward/scrollBackward)
                scroll_command = f'new UiScrollable(new UiSelector().scrollable(true)).scrollForward()'
                scroll_down_and_find_element(AppiumBy.ANDROID_UIAUTOMATOR, scroll_command)
                print(f"Scrolled down to find element")
            except Exception as swipe_error:
                print(f"Swipe failed: {str(swipe_error)}")

            attempts += 1

    print(f"Failed to find element: {value} after {max_attempts} attempts.")
    raise Exception(f"Failed to find element: {value} after {max_attempts} attempts.")


def swipe_left(driver, attempts=3):
    """Perform horizontal scrolling with retries."""
    for attempt in range(attempts):
        try:
            swipe_left_command = 'new UiScrollable(new UiSelector().scrollable(true)).setAsHorizontalList().scrollForward()'
            scroll_down_and_find_element(AppiumBy.ANDROID_UIAUTOMATOR, swipe_left_command)
            logging.info("Scrolled left successfully")
            return True
        except Exception as swipe_error:
            logging.warning(f"Swipe failed on attempt {attempt + 1}: {str(swipe_error)}")
    logging.error("Swipe action failed after retries")
    return False






# Initiate Testing
try:
    # Start Appium session
    driver = webdriver.Remote(appium_server_url, options=UiAutomator2Options().load_capabilities(capabilities))
    logging.info("Session Started")

    # Load Excel data
    SHEET_NAME = 'OnBoarding'
    START_ROW = 0
    MAX_ROW_LIMIT = 20
    validate_elements_from_excel(driver, EXCEL_FILE_PATH, SHEET_NAME, start_row=START_ROW, max_row=MAX_ROW_LIMIT)


    # Perform Swipe Action
    swipe_left(driver)

    # Load Excel data
    SHEET_NAME = 'OnBoarding'
    START_ROW = 21
    MAX_ROW_LIMIT = 29
    validate_elements_from_excel(driver, EXCEL_FILE_PATH, SHEET_NAME, start_row=START_ROW, max_row=MAX_ROW_LIMIT)

    # Login flow
    try:
        scroll_down_and_find_element(AppiumBy.XPATH, XPATH_LOGIN_BUTTON).click()
        logging.info("Login button tapped")

        # Load Excel data
        SHEET_NAME = 'OnBoarding'
        START_ROW = 30
        MAX_ROW_LIMIT = 37
        validate_elements_from_excel(driver, EXCEL_FILE_PATH, SHEET_NAME, start_row=START_ROW, max_row=MAX_ROW_LIMIT)

        # Enter Username
        user_ID = simpledialog.askstring("Hold On!", "Enter your User ID:")
        scroll_down_and_find_element(AppiumBy.XPATH, XPATH_USERNAME_FIELD).send_keys(user_ID)
        logging.info(user_ID)

        # Secure password input
        user_Password = simpledialog.askstring("Hold On!", "Enter your Password:")
        scroll_down_and_find_element(AppiumBy.XPATH, XPATH_PASSWORD_FIELD).send_keys(user_Password)
        logging.info(user_Password)

        # Unmask password
        scroll_down_and_find_element(AppiumBy.XPATH, XPATH_UNMASK_PASSWORD_BUTTON).click()
        logging.info("Unmasked password")

        # Confirm login
        scroll_down_and_find_element(AppiumBy.XPATH, XPATH_LOGIN_BUTTON).click()
        logging.info("Login submitted")

    except Exception as login_error:
        logging.error(f"Login flow failed: {str(login_error)}")

    # Handle OTP and OAC Authentication
    try:
        # OAC Authentication
        oac_auth = simpledialog.askstring("Hold On!", "Enter your OAC:")
        scroll_down_and_find_element(AppiumBy.XPATH, XPATH_OAC_FIELD).send_keys(oac_auth)
        logging.info("OAC entered")
        scroll_down_and_find_element(AppiumBy.XPATH, XPATH_CONFIRM_BUTTON).click()

        # OTP Authentication 
        otp_auth = simpledialog.askstring("Hold On!", "Enter your OTP:")
        scroll_down_and_find_element(AppiumBy.XPATH, XPATH_OTP_FIELD).send_keys(otp_auth)
        logging.info("OTP entered")
        scroll_down_and_find_element(AppiumBy.XPATH, XPATH_CONFIRM_BUTTON).click()

    except Exception as auth_error:
        logging.error(f"Authentication failed: {str(auth_error)}")

finally:
    # Ensure session cleanup
    if driver:
        driver.quit()
        logging.info("Session closed")